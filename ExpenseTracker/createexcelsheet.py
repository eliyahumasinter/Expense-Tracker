from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
def createSheet(header, footer, data, title):
    '''
    
    Title1 Title2 Title3
    --------------------

    '''
    wb = Workbook()
    ws = wb.active
    ws.append(header) #title
    for row in range(2,len(data)+2):
        for col in range(1,len(data[row-2])+1):
            char = get_column_letter(col)
            ws[char + str(row)] = data[row-2][col-1]

    filename = title+'.xlsx'
    ws.append(footer)
    wb.save(filename)
    
    return filename
